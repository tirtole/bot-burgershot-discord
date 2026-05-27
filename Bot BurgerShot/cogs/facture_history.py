from pathlib import Path
from datetime import datetime
import re
import unicodedata

import discord
from discord import app_commands
from discord.ext import commands
from openpyxl import load_workbook

from utils.logger import log_action
from utils.permissions import is_staff


FACTURES_DIR = Path("factures")


def normalize_facture_number(numero: str):
    digits = "".join(char for char in str(numero) if char.isdigit())
    return digits.zfill(4) if digits else ""


def normalize_text(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def clean_text(value, default="Non renseigné"):
    if value is None:
        return default

    text = str(value).strip()

    if not text:
        return default

    return text.replace("`", "'")


def format_cell_value(value):
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).replace(".", ",")

    return str(value).strip()


def format_euro(value):
    try:
        number = float(value)

        if number.is_integer():
            return f"{int(number):,}".replace(",", " ") + "€"

        return f"{number:,.2f}".replace(",", " ").replace(".", ",") + "€"

    except Exception:
        text = str(value or "").strip()

        if not text:
            return "0€"

        if "€" in text:
            return text

        return text


def get_facture_files():
    FACTURES_DIR.mkdir(parents=True, exist_ok=True)

    files = []

    for path in FACTURES_DIR.glob("*"):
        if path.suffix.lower() in [".pdf", ".xlsx"]:
            files.append(path)

    return sorted(
        files,
        key=lambda path: path.stat().st_mtime,
        reverse=True
    )


def get_facture_pairs():
    FACTURES_DIR.mkdir(parents=True, exist_ok=True)

    pairs = {}

    for file in get_facture_files():
        stem = file.stem

        if stem not in pairs:
            pairs[stem] = {
                "pdf": None,
                "xlsx": None
            }

        if file.suffix.lower() == ".pdf":
            pairs[stem]["pdf"] = file

        if file.suffix.lower() == ".xlsx":
            pairs[stem]["xlsx"] = file

    return pairs


def find_facture_pair(numero: str):
    normalized = normalize_facture_number(numero)

    if not normalized:
        return None

    pairs = get_facture_pairs()

    for stem, files in pairs.items():
        if normalized in stem:
            return {
                "stem": stem,
                "pdf": files.get("pdf"),
                "xlsx": files.get("xlsx")
            }

    return None


def get_cell_value(worksheet, row, column):
    if row < 1 or column < 1:
        return None

    return worksheet.cell(row=row, column=column).value


def get_first_value_near(worksheet, row, column):
    positions = [
        (row, column + 1),
        (row, column + 2),
        (row + 1, column),
        (row + 1, column + 1),
        (row + 2, column),
        (row + 2, column + 1),
    ]

    for target_row, target_column in positions:
        value = get_cell_value(worksheet, target_row, target_column)

        if value not in [None, ""]:
            return value

    return None


def find_label_value(workbook, labels):
    normalized_labels = [normalize_text(label) for label in labels]

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                cell_norm = normalize_text(cell.value)

                if not cell_norm:
                    continue

                for label in normalized_labels:
                    if label in cell_norm:
                        raw_text = str(cell.value or "")

                        if " le " in raw_text.lower():
                            after_le = raw_text.lower().split(" le ", 1)[-1].strip()

                            if after_le:
                                return after_le

                        found = get_first_value_near(
                            worksheet,
                            cell.row,
                            cell.column
                        )

                        if found not in [None, ""]:
                            return found

    return None


def extract_block_under_label(workbook, label, max_lines=4):
    target_label = normalize_text(label)
    stop_words = {
        "description",
        "quantite",
        "prix a l unite",
        "prix total",
        "notes",
        "total",
        "facture faite le",
        "facture id",
    }

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if normalize_text(cell.value) == target_label:
                    lines = []

                    for offset in range(1, max_lines + 1):
                        value = get_cell_value(
                            worksheet,
                            cell.row + offset,
                            cell.column
                        )

                        if value in [None, ""]:
                            continue

                        value_text = format_cell_value(value)
                        value_norm = normalize_text(value_text)

                        if value_norm in stop_words:
                            break

                        if value_norm:
                            lines.append(clean_text(value_text))

                    return lines

    return []


def detect_items_columns(row):
    columns = {
        "description": None,
        "quantite": None,
        "prix": None,
        "total": None,
    }

    for cell in row:
        value = normalize_text(cell.value)

        if value == "description":
            columns["description"] = cell.column

        elif "quantite" in value:
            columns["quantite"] = cell.column

        elif "prix a l unite" in value or "prix unite" in value:
            columns["prix"] = cell.column

        elif value == "prix total" or "total" in value:
            columns["total"] = cell.column

    return columns


def find_items_table(workbook):
    for worksheet in workbook.worksheets:
        header_row = None
        columns = None

        for row in worksheet.iter_rows():
            row_norm = [normalize_text(cell.value) for cell in row]

            has_description = "description" in row_norm
            has_quantite = any("quantite" in value for value in row_norm)
            has_total = any("prix total" in value or value == "total" for value in row_norm)

            if has_description and has_quantite and has_total:
                header_row = row[0].row
                columns = detect_items_columns(row)
                break

        if header_row is None or columns is None:
            continue

        col_description = columns.get("description")
        col_quantite = columns.get("quantite")
        col_prix = columns.get("prix")
        col_total = columns.get("total")

        if col_description is None:
            continue

        if col_prix is None and col_quantite is not None:
            col_prix = col_quantite + 1

        if col_total is None and col_prix is not None:
            col_total = col_prix + 1

        items = []

        for row_index in range(header_row + 1, header_row + 12):
            description = get_cell_value(worksheet, row_index, col_description)
            description_text = clean_text(format_cell_value(description), default="")
            description_norm = normalize_text(description_text)

            if description_norm in [
                "notes",
                "total",
                "adjustments",
                "ajustements",
                "facture",
                "facture id",
                "facture faite le",
            ]:
                break

            if not description_norm:
                continue

            if "note" in description_norm or description_norm == "total":
                break

            quantite = get_cell_value(worksheet, row_index, col_quantite) if col_quantite else ""
            prix = get_cell_value(worksheet, row_index, col_prix) if col_prix else ""
            total = get_cell_value(worksheet, row_index, col_total) if col_total else ""

            quantite_text = clean_text(format_cell_value(quantite), default="0")
            prix_text = format_euro(prix)
            total_text = format_euro(total)

            if quantite_text in ["", "Non renseigné"]:
                quantite_text = "0"

            if prix_text in ["", "Non renseigné"]:
                prix_text = "0€"

            if total_text in ["", "Non renseigné"]:
                total_text = "0€"

            items.append({
                "description": description_text,
                "quantite": quantite_text,
                "prix": prix_text,
                "total": total_text,
            })

        return items

    return []


def find_note(workbook):
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = normalize_text(cell.value)

                if value == "notes":
                    positions = [
                        (cell.row + 1, cell.column),
                        (cell.row, cell.column + 1),
                        (cell.row + 1, cell.column + 1),
                    ]

                    for target_row, target_column in positions:
                        note = get_cell_value(
                            worksheet,
                            target_row,
                            target_column
                        )

                        note_text = clean_text(format_cell_value(note), default="")

                        if not note_text:
                            continue

                        note_norm = normalize_text(note_text)

                        if note_norm in ["total", "adjustments", "ajustements"]:
                            continue

                        return note_text

    return ""


def find_total(workbook):
    possible_totals = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value_norm = normalize_text(cell.value)

                if value_norm == "total":
                    found = get_first_value_near(
                        worksheet,
                        cell.row,
                        cell.column
                    )

                    if found not in [None, ""]:
                        possible_totals.append(found)

    for value in possible_totals:
        if isinstance(value, (int, float)):
            return value

    if possible_totals:
        return possible_totals[0]

    return None


def extract_facture_info(xlsx_path: Path | None, pdf_path: Path | None):
    info = {
        "numero_facture": "Inconnu",
        "date_facture": "Inconnue",
        "date_creation": "Inconnue",
        "facture_pour": [],
        "payable_a": [],
        "items": [],
        "note": "",
        "total": None,
        "pdf_path": pdf_path,
        "xlsx_path": xlsx_path,
        "source": "filename"
    }

    if pdf_path is not None and pdf_path.exists():
        modified = datetime.fromtimestamp(pdf_path.stat().st_mtime)
        info["date_creation"] = modified.strftime("%d/%m/%Y %H:%M")

    if xlsx_path is None or not xlsx_path.exists():
        return info

    try:
        workbook = load_workbook(xlsx_path, data_only=True)
    except Exception:
        return info

    info["source"] = "xlsx"

    facture_id = find_label_value(
        workbook,
        ["Facture ID", "Invoice ID"]
    )

    if facture_id:
        info["numero_facture"] = clean_text(format_cell_value(facture_id))

    date_facture = find_label_value(
        workbook,
        ["Facture faite le", "Donnée le", "Donnee le"]
    )

    if date_facture:
        info["date_facture"] = clean_text(format_cell_value(date_facture))

    info["facture_pour"] = extract_block_under_label(
        workbook,
        "Facture pour",
        max_lines=4
    )

    info["payable_a"] = extract_block_under_label(
        workbook,
        "Payable a",
        max_lines=4
    )

    info["items"] = find_items_table(workbook)
    info["note"] = find_note(workbook)
    info["total"] = find_total(workbook)

    return info


def build_items_text(items):
    if not items:
        return "Aucun item trouvé."

    lines = []

    for item in items[:10]:
        lines.append(
            f"• **{item['description']}**\n"
            f"  Quantité : `{item['quantite']}` | "
            f"Prix : `{item['prix']}` | "
            f"Total : `{item['total']}`"
        )

    return "\n".join(lines)


def build_facture_info_embed(info: dict):
    pdf_path = info.get("pdf_path")
    xlsx_path = info.get("xlsx_path")

    pdf_text = f"`{pdf_path.name}`" if pdf_path else "`PDF introuvable`"
    xlsx_text = f"`{xlsx_path.name}`" if xlsx_path else "`XLSX introuvable`"

    facture_pour = "\n".join(info.get("facture_pour") or ["Non renseigné"])
    payable_a = "\n".join(info.get("payable_a") or ["Non renseigné"])

    total = info.get("total")
    total_text = format_euro(total) if total not in [None, ""] else "Inconnu"

    note = clean_text(info.get("note"), default="Aucune note.")

    embed = discord.Embed(
        title="🧾 Facture trouvée",
        color=discord.Color.orange()
    )

    embed.description = (
        "Voici les informations principales récupérées depuis le fichier Excel de la facture."
    )

    embed.add_field(
        name="📌 Informations principales",
        value=(
            f"**Facture ID :** `{info.get('numero_facture', 'Inconnu')}`\n"
            f"**Date facture :** `{info.get('date_facture', 'Inconnue')}`\n"
            f"**Date fichier :** `{info.get('date_creation', 'Inconnue')}`"
        ),
        inline=False
    )

    embed.add_field(
        name="👤 Facture pour",
        value=facture_pour[:1024],
        inline=True
    )

    embed.add_field(
        name="💳 Payable à",
        value=payable_a[:1024],
        inline=True
    )

    embed.add_field(
        name="🍔 Items",
        value=build_items_text(info.get("items", []))[:1024],
        inline=False
    )

    embed.add_field(
        name="💰 Total",
        value=f"`{total_text}`",
        inline=False
    )

    embed.add_field(
        name="📝 Note",
        value=note[:1024],
        inline=False
    )

    embed.add_field(
        name="📁 Fichiers",
        value=(
            f"**PDF :** {pdf_text}\n"
            f"**XLSX :** {xlsx_text}"
        ),
        inline=False
    )

    embed.set_footer(text="BurgerShot Sud RP • Historique factures")
    return embed


class FactureHistory(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    factures_group = app_commands.Group(
        name="factures",
        description="Historique des factures BurgerShot"
    )

    @factures_group.command(
        name="liste",
        description="Afficher les dernières factures"
    )
    async def factures_liste(self, interaction: discord.Interaction):
        if not is_staff(interaction.user):
            await interaction.response.send_message(
                "❌ Tu n’as pas la permission.",
                ephemeral=True
            )
            return

        files = [
            file for file in get_facture_files()
            if file.suffix.lower() == ".pdf"
        ][:15]

        if not files:
            await interaction.response.send_message(
                "❌ Aucune facture trouvée dans le dossier `factures/`.",
                ephemeral=True
            )
            return

        text = ""

        for file in files:
            size_kb = round(file.stat().st_size / 1024, 1)
            modified = datetime.fromtimestamp(file.stat().st_mtime).strftime("%d/%m/%Y %H:%M")
            text += f"• `{file.name}` — `{size_kb} KB` — `{modified}`\n"

        embed = discord.Embed(
            title="🧾 Historique des factures",
            description=(
                f"{text}\n"
                "Utilise `/factures rechercher numero:` pour voir les infos principales.\n"
                "Utilise `/factures renvoyer numero:` pour récupérer une facture."
            ),
            color=discord.Color.orange()
        )

        await interaction.response.send_message(
            embed=embed,
            ephemeral=True
        )

    @factures_group.command(
        name="rechercher",
        description="Rechercher une facture par numéro"
    )
    async def factures_rechercher(
        self,
        interaction: discord.Interaction,
        numero: str
    ):
        if not is_staff(interaction.user):
            await interaction.response.send_message(
                "❌ Tu n’as pas la permission.",
                ephemeral=True
            )
            return

        pair = find_facture_pair(numero)

        if pair is None:
            await interaction.response.send_message(
                f"❌ Aucune facture trouvée pour `{numero}`.",
                ephemeral=True
            )
            return

        info = extract_facture_info(
            xlsx_path=pair.get("xlsx"),
            pdf_path=pair.get("pdf")
        )

        await interaction.response.send_message(
            embed=build_facture_info_embed(info),
            ephemeral=True
        )

    @factures_group.command(
        name="renvoyer",
        description="Renvoyer une facture PDF"
    )
    async def factures_renvoyer(
        self,
        interaction: discord.Interaction,
        numero: str,
        salon: discord.TextChannel | None = None
    ):
        if not is_staff(interaction.user):
            await interaction.response.send_message(
                "❌ Tu n’as pas la permission.",
                ephemeral=True
            )
            return

        pair = find_facture_pair(numero)

        if pair is None or pair.get("pdf") is None:
            await interaction.response.send_message(
                f"❌ Aucune facture PDF trouvée pour `{numero}`.",
                ephemeral=True
            )
            return

        file = pair["pdf"]

        await interaction.response.defer(ephemeral=True)

        facture_file = discord.File(
            str(file),
            filename=file.name
        )

        if salon is None:
            await interaction.followup.send(
                content=f"📎 Facture retrouvée : `{file.name}`",
                file=facture_file,
                ephemeral=True
            )
        else:
            await salon.send(
                content=f"📎 Facture renvoyée par {interaction.user.mention} : `{file.name}`",
                file=facture_file
            )

            await interaction.followup.send(
                f"✅ Facture envoyée dans {salon.mention}.",
                ephemeral=True
            )

        await log_action(
            interaction.guild,
            "🧾 Facture renvoyée",
            f"**Staff :** {interaction.user.mention}\n"
            f"**Facture :** `{file.name}`\n"
            f"**Salon :** {salon.mention if salon else 'Éphémère'}",
            discord.Color.orange()
        )


async def setup(bot):
    await bot.add_cog(FactureHistory(bot))