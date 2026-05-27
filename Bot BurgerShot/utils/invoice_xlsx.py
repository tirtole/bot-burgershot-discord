import json
import re
import zipfile
import subprocess
from copy import copy
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


TEMPLATE_PATHS = [
    Path("templates/Invoice.xlsx"),
    Path("Invoice.xlsx")
]

OUTPUT_DIR = Path("factures")
COUNTER_FILE = Path("data/factures_counter.json")

MAX_ITEMS = 5

LIBREOFFICE_PATHS = [
    Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
    Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"),
]

EURO_FORMAT = '#,##0.00 €'


def get_libreoffice_path():
    for path in LIBREOFFICE_PATHS:
        if path.exists():
            return path

    raise FileNotFoundError(
        "LibreOffice est introuvable.\n"
        "Installe LibreOffice ou vérifie le chemin de `soffice.exe`.\n\n"
        "Chemin attendu : C:\\Program Files\\LibreOffice\\program\\soffice.exe"
    )


def get_template_path():
    for path in TEMPLATE_PATHS:
        if path.exists():
            if not zipfile.is_zipfile(path):
                raise ValueError(
                    f"Le fichier `{path}` n'est pas un vrai .xlsx valide.\n"
                    "Ouvre ton modèle dans Excel/LibreOffice puis fais "
                    "`Enregistrer sous` → `.xlsx`."
                )

            return path

    raise FileNotFoundError(
        "Template Invoice.xlsx introuvable. Mets le fichier dans `templates/Invoice.xlsx`."
    )


def get_writable_cell(cell):
    worksheet = cell.parent

    if not isinstance(cell, MergedCell):
        return cell

    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return worksheet.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            )

    return cell


def safe_set_value(cell, value):
    writable_cell = get_writable_cell(cell)

    try:
        writable_cell.value = value
    except AttributeError:
        pass


def safe_set_number_format(cell, number_format):
    writable_cell = get_writable_cell(cell)

    try:
        writable_cell.number_format = number_format
    except AttributeError:
        pass


def safe_wrap_cell(cell):
    writable_cell = get_writable_cell(cell)

    try:
        alignment = copy(writable_cell.alignment)
        alignment.wrap_text = True
        alignment.vertical = "center"
        writable_cell.alignment = alignment
    except Exception:
        pass


def load_counter():
    COUNTER_FILE.parent.mkdir(parents=True, exist_ok=True)

    if not COUNTER_FILE.exists():
        with open(COUNTER_FILE, "w", encoding="utf-8") as f:
            json.dump({"last_number": 0}, f, indent=4)

    with open(COUNTER_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_counter(data):
    COUNTER_FILE.parent.mkdir(parents=True, exist_ok=True)

    with open(COUNTER_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)


def get_next_invoice_number():
    data = load_counter()
    last_number = int(data.get("last_number", 0))
    next_number = last_number + 1

    data["last_number"] = next_number
    save_counter(data)

    return f"#{next_number:04d}"


def clean_filename(text: str):
    text = str(text).strip()
    text = re.sub(r"[^\w\s-]", "", text, flags=re.UNICODE)
    text = re.sub(r"\s+", "_", text)
    return text[:50] or "facture"


def clean_number(value: str):
    value = str(value).strip()
    value = value.replace("€", "")
    value = value.replace("$", "")
    value = value.replace(" ", "")
    value = value.replace(",", ".")

    number = float(value)

    if number.is_integer():
        return int(number)

    return number


def format_euro(value):
    try:
        number = float(value)

        if number.is_integer():
            return f"{int(number)}€"

        return f"{number:.2f}€"
    except Exception:
        return f"{value}€"


def validate_items(items: list[dict]):
    if not items:
        raise ValueError("Tu dois ajouter au moins 1 item.")

    if len(items) > MAX_ITEMS:
        raise ValueError(f"Tu peux mettre maximum {MAX_ITEMS} items.")

    cleaned_items = []

    for item in items:
        nom = str(item.get("nom", "")).strip()
        quantite = clean_number(item.get("quantite", "0"))
        prix = clean_number(item.get("prix", "0"))

        if not nom:
            raise ValueError("Un item ne peut pas avoir un nom vide.")

        if quantite <= 0:
            raise ValueError(f"La quantité de `{nom}` doit être supérieure à 0.")

        if prix < 0:
            raise ValueError(f"Le prix de `{nom}` ne peut pas être négatif.")

        total = quantite * prix

        cleaned_items.append({
            "nom": nom,
            "quantite": quantite,
            "prix": prix,
            "total": total
        })

    return cleaned_items


def find_tag_cells(workbook):
    tag_cells = {}

    for worksheet in workbook.worksheets:
        already_done = set()

        for row in worksheet.iter_rows():
            for cell in row:
                writable_cell = get_writable_cell(cell)
                key = writable_cell.coordinate

                if key in already_done:
                    continue

                already_done.add(key)

                if not isinstance(writable_cell.value, str):
                    continue

                tags = re.findall(r"<<[^<>]+>>", writable_cell.value)

                for tag in tags:
                    tag_cells.setdefault(tag, []).append(writable_cell)

    return tag_cells


def replace_tags_in_workbook(workbook, mapping: dict):
    for worksheet in workbook.worksheets:
        already_done = set()

        for row in worksheet.iter_rows():
            for cell in row:
                writable_cell = get_writable_cell(cell)
                key = writable_cell.coordinate

                if key in already_done:
                    continue

                already_done.add(key)

                if not isinstance(writable_cell.value, str):
                    continue

                value = writable_cell.value

                for tag, replacement in mapping.items():
                    value = value.replace(tag, str(replacement))

                value = re.sub(r"<<[^<>]+>>", "", value)
                safe_set_value(writable_cell, value)


def only_digits(value):
    return re.sub(r"\D", "", str(value or ""))

def fix_phone_number_format(workbook):
    """
    Corrige le numéro BurgerShot en haut du template :
    75010017 affiché en 75 010 017,00 € -> 075010017
    """
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(min_row=1, max_row=12):
            for cell in row:
                writable_cell = get_writable_cell(cell)

                if writable_cell.value is None:
                    continue

                value_digits = re.sub(r"\D", "", str(writable_cell.value))

                if value_digits in {"75010017", "7501001700"}:
                    safe_set_value(writable_cell, "075010017")
                    writable_cell.number_format = "@"

def style_invoice_title(workbook):
    """
    Corrige juste le titre Facture sans casser le template.
    """
    for worksheet in workbook.worksheets:
        already_done = set()

        for row in worksheet.iter_rows():
            for cell in row:
                writable_cell = get_writable_cell(cell)
                key = writable_cell.coordinate

                if key in already_done:
                    continue

                already_done.add(key)

                if not isinstance(writable_cell.value, str):
                    continue

                if writable_cell.value.strip().lower() == "facture":
                    font = copy(writable_cell.font)
                    font.size = 28
                    font.bold = True
                    writable_cell.font = font

                    worksheet.row_dimensions[writable_cell.row].height = 42

                    alignment = copy(writable_cell.alignment)
                    alignment.vertical = "center"
                    alignment.wrap_text = False
                    writable_cell.alignment = alignment


def set_item_totals_near_price_tags(tag_cells: dict, cleaned_items: list[dict]):
    """
    Remplit le prix total sur la cellule à droite du prix unitaire.
    """
    for index in range(1, MAX_ITEMS + 1):
        price_tag = f"<<prix{index}>>"
        cells = tag_cells.get(price_tag, [])

        for cell in cells:
            worksheet = cell.parent
            total_cell = worksheet.cell(row=cell.row, column=cell.column + 1)

            if index <= len(cleaned_items):
                safe_set_value(total_cell, cleaned_items[index - 1]["total"])
                safe_set_number_format(total_cell, EURO_FORMAT)
            else:
                safe_set_value(total_cell, "")


def keep_empty_item_rows_clean(tag_cells: dict, cleaned_items: list[dict]):
    """
    Garde les lignes grises vides du template.
    Ne cache plus les lignes.
    """
    for index in range(len(cleaned_items) + 1, MAX_ITEMS + 1):
        for tag_name in [
            f"<<item{index}>>",
            f"<<quant{index}>>",
            f"<<prix{index}>>",
            f"<<total{index}>>",
            f"<<total_item{index}>>"
        ]:
            cells = tag_cells.get(tag_name, [])

            for cell in cells:
                safe_set_value(cell, "")


def set_total_near_labels(workbook, total_facture):
    """
    Remplit uniquement la ligne Total.
    """
    for worksheet in workbook.worksheets:
        already_done = set()

        for row in worksheet.iter_rows():
            for cell in row:
                writable_cell = get_writable_cell(cell)
                key = writable_cell.coordinate

                if key in already_done:
                    continue

                already_done.add(key)

                if not isinstance(writable_cell.value, str):
                    continue

                clean_value = writable_cell.value.strip().lower()

                if clean_value == "total":
                    target = worksheet.cell(
                        row=writable_cell.row,
                        column=writable_cell.column + 1
                    )
                    safe_set_value(target, total_facture)
                    safe_set_number_format(target, EURO_FORMAT)


def set_big_bottom_total(workbook, total_facture):
    """
    Remet le gros total rose du bas si le template l'a.
    On cherche une grande cellule en bas avec un format monnaie ou une grande police.
    """
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(min_row=24):
            for cell in row:
                writable_cell = get_writable_cell(cell)

                try:
                    font_size = writable_cell.font.sz or 11
                except Exception:
                    font_size = 11

                number_format = str(writable_cell.number_format or "")

                is_big_total_cell = font_size >= 18
                is_money_cell = "€" in number_format or "$" in number_format

                if is_big_total_cell or is_money_cell:
                    if writable_cell.row >= 25 and writable_cell.column >= 4:
                        safe_set_value(writable_cell, total_facture)
                        safe_set_number_format(writable_cell, EURO_FORMAT)
                        return


def remove_adjustments_only(workbook):
    """
    Supprime seulement Adjustments / Ajustements.
    Ne touche pas au gros total rose.
    """
    adjustment_words = {
        "adjustments",
        "adjustment",
        "ajustements",
        "ajustement"
    }

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                writable_cell = get_writable_cell(cell)

                if not isinstance(writable_cell.value, str):
                    continue

                clean_value = writable_cell.value.strip().lower()

                if clean_value in adjustment_words:
                    safe_set_value(writable_cell, "")

                    target = worksheet.cell(
                        row=writable_cell.row,
                        column=writable_cell.column + 1
                    )
                    safe_set_value(target, "")


def ensure_note_visible(workbook, note: str):
    """
    Si le tag <<note>> n'a pas bien été remplacé à cause d'un template bizarre,
    on force la note sur la ligne juste sous 'Notes:'.
    """
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                writable_cell = get_writable_cell(cell)

                if not isinstance(writable_cell.value, str):
                    continue

                clean_value = writable_cell.value.strip().lower().replace(":", "")

                if clean_value == "notes":
                    target = worksheet.cell(
                        row=writable_cell.row + 1,
                        column=writable_cell.column
                    )
                    safe_set_value(target, note)
                    return


def fix_text_encoding(workbook):
    replacements = {
        "Donn�e": "Donnée",
        "Quantit�": "Quantité",
        "Prix � l'unit�": "Prix à l'unité",
        "Prix à l'unit": "Prix à l'unité",
        "Prix à l'unitée": "Prix à l'unité",
        "$": "€"
    }

    for worksheet in workbook.worksheets:
        already_done = set()

        for row in worksheet.iter_rows():
            for cell in row:
                writable_cell = get_writable_cell(cell)
                key = writable_cell.coordinate

                if key in already_done:
                    continue

                already_done.add(key)

                if isinstance(writable_cell.value, str):
                    value = writable_cell.value

                    for old, new in replacements.items():
                        value = value.replace(old, new)

                    safe_set_value(writable_cell, value)


def apply_invoice_layout_fixes(workbook):
    """
    Ajustements doux : on garde le design du template.
    """
    for worksheet in workbook.worksheets:
        worksheet.sheet_view.showGridLines = False

        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 1
        worksheet.page_setup.orientation = "portrait"
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4

        worksheet.print_area = "A1:G32"

        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.20
        worksheet.page_margins.bottom = 0.20

        worksheet.print_options.horizontalCentered = True
        worksheet.print_options.verticalCentered = False

        worksheet.column_dimensions["B"].width = 30
        worksheet.column_dimensions["C"].width = 18
        worksheet.column_dimensions["D"].width = 18
        worksheet.column_dimensions["E"].width = 15
        worksheet.column_dimensions["F"].width = 17
        worksheet.column_dimensions["G"].width = 20

        for row in [5, 6, 7]:
            worksheet.row_dimensions[row].height = max(
                worksheet.row_dimensions[row].height or 20,
                28
            )

        for row in [18, 19, 20, 21, 22, 23]:
            worksheet.row_dimensions[row].height = max(
                worksheet.row_dimensions[row].height or 20,
                22
            )

        already_done = set()

        for row in worksheet.iter_rows():
            for cell in row:
                writable_cell = get_writable_cell(cell)
                key = writable_cell.coordinate

                if key in already_done:
                    continue

                already_done.add(key)

                if writable_cell.value is None:
                    continue

                if isinstance(writable_cell.value, (int, float)):
                    writable_cell.number_format = EURO_FORMAT

                safe_wrap_cell(writable_cell)


def convert_xlsx_to_pdf(xlsx_path: Path):
    libreoffice = get_libreoffice_path()
    output_dir = xlsx_path.parent
    pdf_path = xlsx_path.with_suffix(".pdf")

    if pdf_path.exists():
        pdf_path.unlink()

    command = [
        str(libreoffice),
        "--headless",
        "--nologo",
        "--nofirststartwizard",
        "--convert-to",
        "pdf",
        "--outdir",
        str(output_dir),
        str(xlsx_path)
    ]

    result = subprocess.run(
        command,
        capture_output=True,
        text=True
    )

    if not pdf_path.exists():
        raise RuntimeError(
            "Impossible de convertir la facture en PDF.\n"
            "Vérifie que LibreOffice est installé.\n\n"
            f"Commande : {' '.join(command)}\n"
            f"STDOUT : {result.stdout}\n"
            f"STDERR : {result.stderr}"
        )

    return pdf_path


def create_invoice_xlsx(
    entreprise: str,
    adresse_entreprise: str,
    mail_entreprise: str,
    destinataire: str,
    numero: str,
    items: list[dict],
    note: str | None = None
):
    template_path = get_template_path()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    cleaned_items = validate_items(items)

    numero_facture = get_next_invoice_number()
    date_facture = datetime.now().strftime("%d/%m/%Y")

    if note is None or not str(note).strip():
        note = "Merci pour votre achat chez BurgerShot."

    total_facture = sum(item["total"] for item in cleaned_items)

    workbook = load_workbook(template_path)
    tag_cells = find_tag_cells(workbook)

    mapping = {
        "<<entreprise>>": entreprise,
        "<<adresse_entreprise>>": adresse_entreprise,
        "<<mail_entreprise>>": mail_entreprise,
        "<<destinataire>>": destinataire,
        "<<numero>>": numero,
        "<<num_facture>>": numero_facture,
        "<<date>>": date_facture,
        "<<Date>>": date_facture,
        "<<note>>": note,
        "<<total>>": total_facture,
        "<<total_facture>>": total_facture,
    }

    for index in range(1, MAX_ITEMS + 1):
        mapping[f"<<item{index}>>"] = ""
        mapping[f"<<quant{index}>>"] = ""
        mapping[f"<<prix{index}>>"] = ""
        mapping[f"<<total{index}>>"] = ""
        mapping[f"<<total_item{index}>>"] = ""

    for index, item in enumerate(cleaned_items, start=1):
        mapping[f"<<item{index}>>"] = item["nom"]
        mapping[f"<<quant{index}>>"] = item["quantite"]
        mapping[f"<<prix{index}>>"] = item["prix"]
        mapping[f"<<total{index}>>"] = item["total"]
        mapping[f"<<total_item{index}>>"] = item["total"]

    replace_tags_in_workbook(workbook, mapping)
    set_item_totals_near_price_tags(tag_cells, cleaned_items)
    keep_empty_item_rows_clean(tag_cells, cleaned_items)
    set_total_near_labels(workbook, total_facture)
    set_big_bottom_total(workbook, total_facture)
    remove_adjustments_only(workbook)
    ensure_note_visible(workbook, note)
    fix_phone_number_format(workbook)
    fix_text_encoding(workbook)
    style_invoice_title(workbook)
    apply_invoice_layout_fixes(workbook)

    safe_destinataire = clean_filename(destinataire)
    safe_numero = numero_facture.replace("#", "")

    xlsx_path = OUTPUT_DIR / f"Facture_{safe_numero}_{safe_destinataire}.xlsx"
    workbook.save(xlsx_path)

    pdf_path = convert_xlsx_to_pdf(xlsx_path)

    return {
        "path": xlsx_path,
        "pdf_path": pdf_path,
        "numero_facture": numero_facture,
        "date": date_facture,
        "entreprise": entreprise,
        "adresse_entreprise": adresse_entreprise,
        "mail_entreprise": mail_entreprise,
        "destinataire": destinataire,
        "numero": numero,
        "items": cleaned_items,
        "note": note,
        "total": total_facture
    }